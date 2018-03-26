# -*- coding=utf-8 -*-
import re
from openpyxl import load_workbook
excel = load_workbook('D:\\gitee\\YunYouKeJiFeiGongKaiXiangMu\\DataProcess\\180323.xlsx')
#获取sheet：
table = excel.get_sheet_by_name('Sheet1')   #通过表名获取
#获取行数和列数：
rows = table.max_row   #获取行数
cols = table.max_column    #获取列数
print(rows, cols)
#获取单元格值：
for i in range(1,rows+1):
    Data = table.cell(row=i, column=1).value  # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
    #print(Data)
    reresult = re.search('mem', str(Data))
    if reresult != None:
        reresult = reresult.span()
    #print(reresult)

    if reresult != None:
        table.cell(row=i, column=1, value='')
    elif Data == None:
        table.cell(row=i, column=1, value='')
    elif Data == 'heart beat':
        table.cell(row=i-6, column=1, value='')
        table.cell(row=i-5, column=1, value='')
        table.cell(row=i-4, column=1, value='')
        table.cell(row=i-3, column=1, value='')
        table.cell(row=i-2, column=1, value='')
        table.cell(row=i-1, column=1, value='')
        table.cell(row=i, column=1, value='')

excel.save('D:\\gitee\\YunYouKeJiFeiGongKaiXiangMu\\DataProcess\\balances180323.xlsx')